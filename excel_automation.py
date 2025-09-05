import openpyxl
from openpyxl import Workbook
from datetime import datetime, timedelta
import calendar
import os

def update_excel_calendar(file_path):
    """Update Excel file with automatic date and day formulas"""
    
    # Check if file exists
    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
    else:
        wb = Workbook()
    
    # Get or create the first worksheet
    if wb.worksheets:
        ws = wb.active
    else:
        ws = wb.create_sheet("Calendar")
    
    # Set up headers
    ws['A1'] = '年 (Year)'
    ws['B1'] = 2025
    ws['A2'] = '月 (Month)'
    ws['B2'] = 1
    
    # Headers for calendar
    ws['A4'] = '日付 (Date)'
    ws['B4'] = '曜日 (Day)'
    ws['C4'] = '売上 (Sales)'
    ws['D4'] = 'メモ (Notes)'
    
    # Add formulas for automatic date calculation
    for row in range(5, 36):  # Up to 31 days
        day_num = row - 4
        
        # Date formula - creates date from year, month, and day
        date_formula = f'=IF({day_num}<=DAY(EOMONTH(DATE($B$1,$B$2,1),0)),DATE($B$1,$B$2,{day_num}),"")'
        ws[f'A{row}'] = date_formula
        
        # Day of week formula
        day_formula = f'=IF(A{row}<>"",TEXT(A{row},"dddd"),"")'
        ws[f'B{row}'] = day_formula
        
        # Format the date cell
        ws[f'A{row}'].number_format = 'yyyy/mm/dd'
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    
    # Save the workbook
    wb.save(file_path)
    print(f"Updated: {file_path}")

# Update both Excel files
excel_dir = r"C:\Users\user\OneDrive\デスクトップ\売上カレンダー"
file1 = os.path.join(excel_dir, "_____.xlsx")
file2 = os.path.join(excel_dir, "_________2025.xlsx")

print("Updating Excel files with automatic date formulas...")
update_excel_calendar(file1)
update_excel_calendar(file2)
print("Excel files updated successfully!")